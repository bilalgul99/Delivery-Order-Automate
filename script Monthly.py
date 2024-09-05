import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import openpyxl
# Load the database (pallet data) file
pallet_data = pd.read_excel('Pallet data.xlsx', usecols=[1, 9, 11], header=None, names=['SKU', 'PalletQty', 'PalletWeight'])

# Load the input Excel file
input_file = 'input.xlsx'
xls = pd.ExcelFile(input_file)

max_weight = 40000
max_pallets = 45

def process_sheet(sheet):
    df = pd.read_excel(xls, sheet_name=sheet, header=None)

    # Check if the second row contains the headers
    headers = df.iloc[1].values

    # Find the pattern 'SKU', 'Volume' (optional), 'Qty'
    pattern = ['SKU', 'Volume', 'Qty']
    pattern_without_volume = ['SKU', 'Qty']

    # Initialize variables to track the number of tables and their positions
    num_tables = 0
    table_positions = []

    # Iterate over the headers to find the pattern
    for i in range(len(headers)):
        if headers[i] == 'SKU':
            if i + 2 < len(headers) and headers[i + 1] == 'Volume' and headers[i + 2] == 'Qty':
                num_tables += 1
                table_positions.append(i)
            elif i + 1 < len(headers) and headers[i + 1] == 'Qty':
                num_tables += 1
                table_positions.append(i)

    # Determine the number of columns in each table
    num_columns = 3 if 'Volume' in headers else 2

    # Create a new dataframe for output
    output_df = pd.DataFrame()

    # Process each table
    for table_index, table_pos in enumerate(table_positions):
        # Extract the table data
        table_data = df.iloc[:, table_pos:table_pos + num_columns]

        # Process the current table
        current_order = {'weight': 0, 'pallets': 0, 'data': []}
        order_index = 1
        orders = []

        # Process each row in the table
        for i, row in table_data.iterrows():
            if row.iloc[0] == 'SKU':
                # Close and save the previous order if it has data
                if current_order['data']:
                    orders.append(current_order)
                    current_order = {'weight': 0, 'pallets': 0, 'data': []}
                    order_index = 1
                continue

            sku = row.iloc[0]
            qty = row.iloc[-1]

            # Find the pallet quantity and weight
            pallet_info = pallet_data[pallet_data['SKU'] == sku]
            if not pallet_info.empty:
                pallet_qty = pallet_info.iloc[0]['PalletQty']
                pallet_weight = pallet_info.iloc[0]['PalletWeight']

                # Calculate the number of pallets to ship
                ship_qty = (qty // pallet_qty) * pallet_qty
                if ship_qty == 0:
                    continue

                pallets = ship_qty // pallet_qty
                weight = pallets * pallet_weight

                # Check if the current order exceeds limits
                if (current_order['weight'] + weight > max_weight) or (current_order['pallets'] + pallets > max_pallets):
                    # Calculate how many pallets can fit in the current order
                    weight_space = max_weight - current_order['weight']
                    pallet_space = max_pallets - current_order['pallets']
                    
                    pallets_to_add = min(pallets, pallet_space, weight_space // pallet_weight)
                    
                    if pallets_to_add > 0:
                        # Add as many pallets as possible to the current order
                        current_order['weight'] += pallets_to_add * pallet_weight
                        current_order['pallets'] += pallets_to_add
                        current_order['data'].append((sku, pallets_to_add * pallet_qty, pallets_to_add, order_index, i))
                    
                    # Finalize the current order
                    orders.append(current_order)
                    
                    # Start a new order with the remaining pallets
                    remaining_pallets = pallets - pallets_to_add
                    order_index += 1
                    current_order = {
                        'weight': remaining_pallets * pallet_weight,
                        'pallets': remaining_pallets,
                        'data': [(sku, remaining_pallets * pallet_qty, remaining_pallets, order_index, i)] if remaining_pallets > 0 else []
                    }
                else:
                    # Update the current order
                    current_order['weight'] += weight
                    current_order['pallets'] += pallets
                    current_order['data'].append((sku, ship_qty, pallets, order_index, i))

        # Append the last order if it has data
        if current_order['data']:
            orders.append(current_order)

        # Create output columns for the current table
        max_orders = max(order['data'][-1][3] for order in orders) if orders else 0
        output_columns = pd.DataFrame(index=table_data.index)
        
        for i in range(1, max_orders + 1):
            output_columns[f'Order {i} Qty'] = ''
            output_columns[f'Order {i} Pallets'] = ''

        # Fill the output columns
        for order in orders:
            for item in order['data']:
                sku, ship_qty, pallets, index, row_num = item
                output_columns.loc[row_num, f'Order {index} Qty'] = ship_qty
                output_columns.loc[row_num, f'Order {index} Pallets'] = pallets

        # Add order headers to the 'SKU' row
        sku_row_index = table_data.index[table_data.iloc[:, 0] == 'SKU'][0]
        for i in range(1, max_orders + 1):
            output_columns.loc[sku_row_index, f'Order {i} Qty'] = f'Order {i}'
            output_columns.loc[sku_row_index, f'Order {i} Pallets'] = 'Pallets'

        # Combine the original table with its output
        combined_table = pd.concat([table_data, output_columns], axis=1)
        
        # Add the combined table to the output dataframe
        output_df = pd.concat([output_df, combined_table], axis=1)

    return output_df, table_positions, num_columns

# Process each sheet in the Excel file
output_sheets = {}
table_info = {}
for sheet in xls.sheet_names:
    output_sheets[sheet], table_positions, num_columns = process_sheet(sheet)
    table_info[sheet] = (table_positions, num_columns)

# Save the result to a new Excel file and apply formatting
with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
    for sheet_name, df in output_sheets.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    # Load the original workbook to copy styles
    original_wb = load_workbook(input_file)
    
    # Get the workbook from the ExcelWriter object
    wb = writer.book

    for sheet_name, df in output_sheets.items():
        ws = wb[sheet_name]
        original_ws = original_wb[sheet_name]
        table_positions, num_columns = table_info[sheet_name]

        # Copy merged cells
        for merged_cell_range in original_ws.merged_cells.ranges:
            ws.merge_cells(str(merged_cell_range))

        # Copy formatting from original sheet
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                original_cell = original_ws.cell(row=cell.row, column=cell.column)
                cell.font = original_cell.font.copy()
                cell.border = original_cell.border.copy()
                cell.fill = original_cell.fill.copy()
                cell.number_format = original_cell.number_format
                cell.alignment = original_cell.alignment.copy()

        # Apply yellow background to output columns
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for table_start in table_positions:
            # Find the max_orders for this table
            max_orders = 0
            for col in range(table_start + num_columns + 1, ws.max_column + 1, 2):
                header = ws.cell(row=2, column=col).value
                if header and header.startswith('Order '):
                    max_orders = max(max_orders, int(header.split()[1]))
                else:
                    break

            start_col = table_start + num_columns + 1
            end_col = start_col + (max_orders * 2)  # 2 columns per order
            for col in range(start_col, end_col):
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = yellow_fill
                    # Preserve borders
                    if cell.border:
                        cell.border = cell.border.copy()

        # Auto-fit column widths
        for col in ws.columns:
            max_length = 0
            column = None
            for cell in col:
                if isinstance(cell, openpyxl.cell.cell.Cell):  # Check if it's a regular cell
                    if column is None:
                        column = cell.column_letter
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            if column:  # Only adjust if we found a valid column
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

    wb.save('output.xlsx')
